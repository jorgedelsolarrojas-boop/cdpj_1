# validator/validators/validacion_report.py

from pathlib import Path
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Image, Table, TableStyle
)
import matplotlib.pyplot as plt
import io
import datetime

# ---------------- utilidades y funciones auxiliares (copiadas de tu script) ----------------

DNI_CANDIDATES = ['DNI','dni','Nº documento','NRO DOCUMENTO','DOCUMENTO','documento','NRO_DOC']
DIAS_CANDIDATES = ['dias_restantes','dias restantes','dias','días','dias_rest','dias_rest']
ESTADO_CANDIDATES = ['estado','Estado','ESTADO']

def find_column(df, candidates):
    cols_lower = {c.lower(): c for c in df.columns}
    for cand in candidates:
        if cand.lower() in cols_lower:
            return cols_lower[cand.lower()]
    for cand in candidates:
        for c in df.columns:
            if cand.lower() in str(c).lower():
                return c
    return None

def load_excel(path: Path) -> pd.DataFrame:
    if not path.exists():
        raise FileNotFoundError(f"Archivo no encontrado: {path}")
    return pd.read_excel(path)

def normalize_dni_series(s: pd.Series) -> pd.Series:
    return (
        s.fillna('')
         .astype(str)
         .str.replace(r"[\.\s-]", '', regex=True)
         .str.upper()
         .str.zfill(8)
    )

def force_dni_text(out_df: pd.DataFrame, dni_col_name: str):
    if dni_col_name in out_df.columns:
        out_df[dni_col_name] = normalize_dni_series(out_df[dni_col_name]).astype(str)
    else:
        out_df.insert(0, dni_col_name, pd.Series([''] * len(out_df)))

def set_xlsx_column_text_format(xlsx_path: Path, dni_col_name: str):
    wb = load_workbook(xlsx_path)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    if dni_col_name in headers:
        col_idx = headers.index(dni_col_name) + 1
        col_letter = get_column_letter(col_idx)
        for cell in ws[col_letter]:
            cell.number_format = '@'
        wb.save(xlsx_path)

def build_maps_and_sets(df_detalles, df_sus, dni_det_col, dni_sus_col, dias_col, estado_col=None):
    df_detalles['_dni_norm'] = normalize_dni_series(df_detalles[dni_det_col])
    df_sus['_dni_norm'] = normalize_dni_series(df_sus[dni_sus_col])

    set_detalles = set(df_detalles['_dni_norm'].dropna().unique())
    set_sus = set(df_sus['_dni_norm'].dropna().unique())

    df_sus['_dias_num'] = pd.to_numeric(df_sus[dias_col], errors='coerce')
    dias_map = df_sus.groupby('_dni_norm')['_dias_num'].first().to_dict()

    descargado_set = set()
    if estado_col and estado_col in df_sus.columns:
        estado_series = df_sus[estado_col].fillna('').astype(str).str.strip().str.lower()
        descargados = df_sus.loc[estado_series == 'descargado', '_dni_norm']
        descargado_set = set(descargados.dropna().unique())

    return set_detalles, set_sus, dias_map, descargado_set

def classify_rows(df_bruto, dni_bruto_col, set_detalles, set_sus, dias_map, descargado_set=None):
    df = df_bruto.copy()
    df['_dni_norm'] = normalize_dni_series(df[dni_bruto_col])

    motivos, passes = [], []
    for dni in df['_dni_norm']:
        motivo, in_det, in_sus = [], dni in set_detalles, dni in set_sus
        dias = dias_map.get(dni, np.nan)
        dias_ok = pd.notna(dias) and dias < 30

        if not in_det: motivo.append('no_en_detalles')
        if not in_sus: motivo.append('no_en_suscriptores')
        if descargado_set and dni in descargado_set: motivo.append('descargado')
        if not pd.notna(dias):
            if in_sus: motivo.append('sin_dias')
        else:
            if not dias_ok: motivo.append('dias>=30')

        motivos.append(';'.join(motivo) if motivo else '')
        passes.append(len(motivo) == 0)

    df['_motivo_rechazo'] = motivos
    df['_passes'] = passes
    return df

def save_outputs(df_validos, df_rechazados, orig_dni_col, out_valid_xlsx, out_rech_xlsx, out_valid_csv, out_rech_csv):
    force_dni_text(df_validos, orig_dni_col)
    force_dni_text(df_rechazados, orig_dni_col)

    dfv_csv, dfr_csv = df_validos.copy(), df_rechazados.copy()
    dfv_csv[orig_dni_col] = dfv_csv[orig_dni_col].apply(lambda x: f"'{x}")
    dfr_csv[orig_dni_col] = dfr_csv[orig_dni_col].apply(lambda x: f"'{x}")

    df_validos.to_excel(out_valid_xlsx, index=False)
    df_rechazados.to_excel(out_rech_xlsx, index=False)
    dfv_csv.to_csv(out_valid_csv, index=False)
    dfr_csv.to_csv(out_rech_csv, index=False)

    set_xlsx_column_text_format(out_valid_xlsx, orig_dni_col)
    set_xlsx_column_text_format(out_rech_xlsx, orig_dni_col)

def create_pie_chart(count_valid, count_reject, output_bytes_io):
    labels = ['Válidos', 'Rechazados']
    sizes = [count_valid, count_reject]
    fig, ax = plt.subplots(figsize=(4,4))
    ax.pie(sizes, labels=labels, autopct='%1.1f%%', startangle=90)
    ax.axis('equal')
    plt.tight_layout()
    plt.savefig(output_bytes_io, format='png', bbox_inches='tight')
    plt.close(fig)
    output_bytes_io.seek(0)

def make_pdf_report(summary, df_bruto, df_validos, df_rechazados, orig_dni_col, out_pdf):
    doc = SimpleDocTemplate(str(out_pdf), pagesize=A4, rightMargin=30,leftMargin=30, topMargin=30,bottomMargin=18)
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph("Reporte de Validación de Registros", styles['Heading1']))
    story.append(Spacer(1, 8))
    story.append(Paragraph(f"Fecha: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
    story.append(Paragraph(f"Total registros: <b>{summary['total']}</b>", styles['Normal']))
    story.append(Paragraph(f"Válidos: <b>{summary['validos']}</b>", styles['Normal']))
    story.append(Paragraph(f"Rechazados: <b>{summary['rechazados']}</b>", styles['Normal']))
    story.append(Spacer(1, 12))

    buf = io.BytesIO()
    create_pie_chart(summary['validos'], summary['rechazados'], buf)
    story.append(Image(buf, width=200, height=200))
    story.append(Spacer(1, 12))

    motivos_counts = summary.get('motivos_counts', {})
    table_data = [['Motivo','Cantidad']] + [[k,str(v)] for k,v in motivos_counts.items()]
    tbl = Table(table_data, colWidths=[250,80])
    tbl.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.lightgrey),
                             ('GRID',(0,0),(-1,-1),0.5,colors.grey)]))
    story.append(Paragraph("Conteo por motivo de rechazo", styles['Heading2']))
    story.append(tbl)

    doc.build(story)

# ---------------- función principal exportable ----------------

def run_validation(path_bruto, path_detalles, path_sus, out_dir=None):
    """
    Ejecuta la validación de registros.
    Retorna un dict con las rutas de salida.
    """
    path_bruto, path_detalles, path_sus = map(Path, [path_bruto, path_detalles, path_sus])
    out_dir = Path(out_dir) if out_dir else path_bruto.parent
    out_dir.mkdir(parents=True, exist_ok=True)

    out_valid_xlsx = out_dir / 'validos.xlsx'
    out_rech_xlsx = out_dir / 'rechazados.xlsx'
    out_valid_csv = out_dir / 'validos.csv'
    out_rech_csv = out_dir / 'rechazados.csv'
    out_pdf = out_dir / 'reporte_validacion.pdf'

    # cargar archivos
    df_bruto, df_detalles, df_sus = map(load_excel, [path_bruto, path_detalles, path_sus])

    # detectar columnas
    dni_bruto_col = find_column(df_bruto, DNI_CANDIDATES)
    dni_det_col = find_column(df_detalles, DNI_CANDIDATES)
    dni_sus_col = find_column(df_sus, DNI_CANDIDATES)
    if not dni_bruto_col or not dni_det_col or not dni_sus_col:
        raise ValueError("No se pudo detectar columna DNI en alguno de los archivos")

    dias_col = find_column(df_sus, DIAS_CANDIDATES)
    if not dias_col:
        raise ValueError("No se pudo detectar columna de días en suscriptores")

    estado_col = find_column(df_sus, ESTADO_CANDIDATES)

    # sets y mapas
    set_detalles, set_sus, dias_map, descargado_set = build_maps_and_sets(
        df_detalles, df_sus, dni_det_col, dni_sus_col, dias_col, estado_col
    )

    df_clas = classify_rows(df_bruto, dni_bruto_col, set_detalles, set_sus, dias_map, descargado_set)
    df_clas['_dias_assoc'] = df_clas['_dni_norm'].map(dias_map)

    df_validos = df_clas[df_clas['_passes']].copy()
    df_rechazados = df_clas[~df_clas['_passes']].copy()

    for c in ['_dni_norm','_passes']:
        if c in df_validos: df_validos.drop(columns=[c], inplace=True)
        if c in df_rechazados and c != '_motivo_rechazo': df_rechazados.drop(columns=[c], inplace=True)

    orig_dni_col = dni_bruto_col
    if orig_dni_col not in df_validos.columns:
        df_validos.insert(0, orig_dni_col, df_clas['_dni_norm'])
    if orig_dni_col not in df_rechazados.columns:
        df_rechazados.insert(0, orig_dni_col, df_clas['_dni_norm'])

    save_outputs(df_validos, df_rechazados, orig_dni_col, out_valid_xlsx, out_rech_xlsx, out_valid_csv, out_rech_csv)

    summary = {
        'total': len(df_clas),
        'validos': len(df_validos),
        'rechazados': len(df_rechazados),
        'motivos_counts': df_clas['_motivo_rechazo'].replace('', np.nan).value_counts().to_dict()
    }

    make_pdf_report(summary, df_clas, df_validos, df_rechazados, orig_dni_col, out_pdf)

    return {
        'validos_xlsx': str(out_valid_xlsx),
        'rechazados_xlsx': str(out_rech_xlsx),
        'validos_csv': str(out_valid_csv),
        'rechazados_csv': str(out_rech_csv),
        'pdf': str(out_pdf),
    }
